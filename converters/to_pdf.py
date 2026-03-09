"""
PDFToolKit - Diğer Formatlardan PDF'e Dönüştürme

Word (DOCX), Excel (XLSX), JPG → PDF dönüştürme fonksiyonları.

Windows'ta Microsoft Office COM otomasyonu ile birebir sayfa düzeni
korunarak dönüşüm yapılır. Office yoksa fpdf2 fallback kullanılır.
"""

import sys
from pathlib import Path

from utils.logger import get_logger
from utils.file_utils import generate_output_filename

logger = get_logger(__name__)


def _office_com_available() -> bool:
    """Windows'ta Office COM otomasyonu kullanılabilir mi kontrol eder."""
    if sys.platform != "win32":
        return False
    try:
        import win32com.client  # noqa: F401
        return True
    except ImportError:
        return False


# ─── Word → PDF ──────────────────────────────────────────────────────────────

def _word_to_pdf_com(docx_path: Path, output_path: Path) -> Path:
    """
    Microsoft Word COM otomasyonu ile DOCX → PDF dönüşümü.
    Sayfa düzeni, fontlar, stiller, tablolar birebir korunur.
    """
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    word = None
    doc = None
    try:
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        word.DisplayAlerts = False

        doc = word.Documents.Open(str(docx_path.resolve()))
        # wdFormatPDF = 17
        doc.SaveAs2(str(output_path.resolve()), FileFormat=17)
        logger.info(f"Word → PDF (COM): '{docx_path.name}' → {output_path}")
    finally:
        if doc:
            doc.Close(False)
        if word:
            word.Quit()
        pythoncom.CoUninitialize()

    return output_path


def _word_to_pdf_fallback(docx_path: Path, output_path: Path) -> Path:
    """
    fpdf2 ile basit DOCX → PDF dönüşümü (fallback).
    Tam düzen korunamaz ancak metin içeriği aktarılır.
    """
    from docx import Document
    from fpdf import FPDF

    doc = Document(str(docx_path))
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Arial TTF fontu ile Unicode/Türkçe karakter desteği
    _register_arial_font(pdf)
    pdf.add_page()
    pdf.set_font("Arial", size=11)

    for paragraph in doc.paragraphs:
        text = paragraph.text
        if not text.strip():
            pdf.ln(5)
            continue

        if paragraph.style.name.startswith("Heading"):
            level = paragraph.style.name.replace("Heading ", "")
            try:
                level = int(level)
                font_size = max(18 - (level * 2), 12)
            except ValueError:
                font_size = 14
            pdf.set_font("Arial", "B", font_size)
            pdf.multi_cell(0, 8, text)
            pdf.set_font("Arial", size=11)
        else:
            pdf.multi_cell(0, 6, text)
        pdf.ln(2)

    pdf.output(str(output_path))
    logger.info(f"Word → PDF (fallback): '{docx_path.name}' → {output_path}")
    return output_path


def word_to_pdf(
    docx_path: str | Path,
    output_path: str | Path | None = None
) -> Path:
    """
    Word (DOCX) dosyasını PDF'e dönüştürür.

    Windows + Microsoft Word kuruluysa COM otomasyonu ile birebir
    sayfa düzeni korunarak dönüşüm yapılır. Aksi halde fpdf2 fallback.

    Args:
        docx_path: Kaynak DOCX dosyası.
        output_path: Çıktı dosya yolu. None ise otomatik üretilir.

    Returns:
        Oluşturulan PDF'nin Path nesnesi.
    """
    docx_path = Path(docx_path)

    if not docx_path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {docx_path}")

    if output_path is None:
        output_path = generate_output_filename(docx_path.stem, "", ".pdf")
    else:
        output_path = Path(output_path)

    if _office_com_available():
        try:
            return _word_to_pdf_com(docx_path, output_path)
        except Exception as e:
            logger.warning(
                f"Word COM dönüşümü başarısız, fallback kullanılıyor: {e}"
            )

    return _word_to_pdf_fallback(docx_path, output_path)


# ─── Excel → PDF ─────────────────────────────────────────────────────────────

def _excel_to_pdf_com(xlsx_path: Path, output_path: Path) -> Path:
    """
    Microsoft Excel COM otomasyonu ile XLSX → PDF dönüşümü.
    Hücre yapısı, sütun genişlikleri, formatlar birebir korunur.
    """
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        # xlTypePDF = 0
        wb.ExportAsFixedFormat(0, str(output_path.resolve()))
        logger.info(f"Excel → PDF (COM): '{xlsx_path.name}' → {output_path}")
    finally:
        if wb:
            wb.Close(False)
        if excel:
            excel.Quit()
        pythoncom.CoUninitialize()

    return output_path


def _excel_to_pdf_fallback(xlsx_path: Path, output_path: Path) -> Path:
    """
    fpdf2 ile basit XLSX → PDF dönüşümü (fallback).
    Tam hücre yapısı korunamaz ancak tablo içeriği aktarılır.
    """
    from openpyxl import load_workbook
    from fpdf import FPDF

    wb = load_workbook(str(xlsx_path), data_only=True)
    pdf = FPDF(orientation="L")
    pdf.set_auto_page_break(auto=True, margin=15)

    _register_arial_font(pdf)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, f"Sayfa: {sheet_name}", ln=True)
        pdf.set_font("Arial", size=9)
        pdf.ln(3)

        max_cols = ws.max_column or 1
        page_width = pdf.w - 2 * pdf.l_margin
        col_width = min(page_width / max_cols, 45)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            for cell_value in row:
                text = str(cell_value) if cell_value is not None else ""
                if len(text) > 25:
                    text = text[:22] + "..."
                pdf.cell(col_width, 7, text, border=1)
            pdf.ln()

    pdf.output(str(output_path))
    logger.info(f"Excel → PDF (fallback): '{xlsx_path.name}' → {output_path}")
    return output_path


def excel_to_pdf(
    xlsx_path: str | Path,
    output_path: str | Path | None = None
) -> Path:
    """
    Excel (XLSX) dosyasını PDF'e dönüştürür.

    Windows + Microsoft Excel kuruluysa COM otomasyonu ile birebir
    hücre yapısı korunarak dönüşüm yapılır. Aksi halde fpdf2 fallback.

    Args:
        xlsx_path: Kaynak XLSX dosyası.
        output_path: Çıktı dosya yolu. None ise otomatik üretilir.

    Returns:
        Oluşturulan PDF'nin Path nesnesi.
    """
    xlsx_path = Path(xlsx_path)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {xlsx_path}")

    if output_path is None:
        output_path = generate_output_filename(xlsx_path.stem, "", ".pdf")
    else:
        output_path = Path(output_path)

    if _office_com_available():
        try:
            return _excel_to_pdf_com(xlsx_path, output_path)
        except Exception as e:
            logger.warning(
                f"Excel COM dönüşümü başarısız, fallback kullanılıyor: {e}"
            )

    return _excel_to_pdf_fallback(xlsx_path, output_path)


# ─── Arial font yardımcıları (fallback için) ─────────────────────────────────

# Windows Arial font dosyaları
_ARIAL_FONT_DIR = Path("C:/Windows/Fonts")
_ARIAL_REGULAR = _ARIAL_FONT_DIR / "arial.ttf"
_ARIAL_BOLD = _ARIAL_FONT_DIR / "arialbd.ttf"
_ARIAL_ITALIC = _ARIAL_FONT_DIR / "ariali.ttf"


def _register_arial_font(pdf) -> None:
    """
    FPDF nesnesine Arial TTF fontunu kaydeder (Unicode/Türkçe karakter desteği).
    """
    if not _ARIAL_REGULAR.exists():
        logger.warning(
            f"Arial font bulunamadi: {_ARIAL_REGULAR} — "
            "Helvetica kullanilacak (Turkce karakterler eksik olabilir)"
        )
        return

    pdf.add_font("Arial", "", str(_ARIAL_REGULAR))
    if _ARIAL_BOLD.exists():
        pdf.add_font("Arial", "B", str(_ARIAL_BOLD))
    if _ARIAL_ITALIC.exists():
        pdf.add_font("Arial", "I", str(_ARIAL_ITALIC))


# ─── JPG → PDF ───────────────────────────────────────────────────────────────

def jpg_to_pdf(
    image_path: str | Path,
    output_path: str | Path | None = None
) -> Path:
    """
    Tek bir JPG/PNG görselini PDF'e dönüştürür.

    Args:
        image_path: Kaynak görsel dosyası.
        output_path: Çıktı dosya yolu. None ise otomatik üretilir.

    Returns:
        Oluşturulan PDF'nin Path nesnesi.
    """
    from PIL import Image

    image_path = Path(image_path)

    if not image_path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {image_path}")

    # Çıktı yolu
    if output_path is None:
        output_path = generate_output_filename(image_path.stem, "", ".pdf")
    else:
        output_path = Path(output_path)

    image = Image.open(str(image_path))

    # RGBA → RGB (PDF RGBA desteklemez)
    if image.mode == "RGBA":
        background = Image.new("RGB", image.size, (255, 255, 255))
        background.paste(image, mask=image.split()[3])
        image = background
    elif image.mode != "RGB":
        image = image.convert("RGB")

    image.save(str(output_path), "PDF", resolution=100.0)

    logger.info(f"JPG → PDF: '{image_path.name}' → {output_path}")
    return output_path


def images_to_pdf(
    image_list: list[str | Path],
    output_path: str | Path | None = None
) -> Path:
    """
    Birden fazla görseli tek bir PDF'de birleştirir.

    Args:
        image_list: Görsel dosya yolları listesi.
        output_path: Çıktı dosya yolu. None ise otomatik üretilir.

    Returns:
        Oluşturulan PDF'nin Path nesnesi.
    """
    from PIL import Image

    if not image_list:
        raise ValueError("Görsel listesi boş")

    # Çıktı yolu
    if output_path is None:
        output_path = generate_output_filename("gorseller", "_birlesik", ".pdf")
    else:
        output_path = Path(output_path)

    images = []
    for img_path in image_list:
        img_path = Path(img_path)
        if not img_path.exists():
            logger.warning(f"Görsel bulunamadı, atlandı: {img_path}")
            continue

        img = Image.open(str(img_path))

        # RGBA → RGB
        if img.mode == "RGBA":
            background = Image.new("RGB", img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")

        images.append(img)

    if not images:
        raise ValueError("Hiçbir geçerli görsel bulunamadı")

    # İlk görsel ana, diğerleri ek sayfa
    first = images[0]
    rest = images[1:] if len(images) > 1 else []

    first.save(
        str(output_path), "PDF",
        resolution=100.0,
        save_all=True,
        append_images=rest
    )

    logger.info(f"{len(images)} görsel → PDF: {output_path}")
    return output_path


# ─── Toplu Dönüştür ve Birleştir ─────────────────────────────────────────────

def convert_and_merge(
    file_paths: list[str | Path],
    converter_func,
    output_path: str | Path | None = None,
    add_page_numbers: bool = False,
    start_number: int = 1,
) -> Path:
    """
    Birden fazla dosyayı önce ayrı ayrı PDF'e dönüştürür,
    ardından tek bir PDF'de birleştirir.

    Args:
        file_paths: Dönüştürülecek dosya yolları (sıralı).
        converter_func: Tek dosya dönüştürme fonksiyonu (path → Path).
        output_path: Çıktı dosya yolu. None ise otomatik üretilir.
        add_page_numbers: Sayfa numaralandırma yapılsın mı.
        start_number: Numaralandırma başlangıç değeri.

    Returns:
        Oluşturulan birleşik PDF'nin Path nesnesi.
    """
    import tempfile
    import shutil
    from core.pdf_merger import merge_pdfs

    if not file_paths:
        raise ValueError("Dosya listesi boş")

    if output_path is None:
        output_path = generate_output_filename("birlesik", "", ".pdf")
    else:
        output_path = Path(output_path)

    # Geçici klasörde ara PDF'leri oluştur
    temp_dir = Path(tempfile.mkdtemp(prefix="pdftoolkit_"))
    try:
        temp_pdfs: list[Path] = []
        for i, fp in enumerate(file_paths):
            fp = Path(fp)
            temp_out = temp_dir / f"{i:04d}_{fp.stem}.pdf"
            converter_func(fp, temp_out)
            temp_pdfs.append(temp_out)
            logger.info(f"Ara dönüşüm ({i+1}/{len(file_paths)}): {fp.name}")

        # Birleştir
        file_list = [{"path": str(p), "pages": None} for p in temp_pdfs]
        merge_pdfs(file_list, output_path)

        # Numaralandırma
        if add_page_numbers:
            _stamp_all_pages(output_path, start_number)

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    logger.info(
        f"{len(file_paths)} dosya → tek PDF: {output_path}"
        f"{' (numaralı)' if add_page_numbers else ''}"
    )
    return output_path


def _stamp_all_pages(
    pdf_path: Path,
    start: int = 1,
    mode: str = "sequential",
    page_counts: list[int] | None = None,
) -> None:
    """PDF'deki her sayfaya sayfa numarası ekler.

    Args:
        pdf_path: Numaralandırılacak PDF dosyası.
        start: Başlangıç numarası.
        mode: "sequential" (ardışık) veya "per_document" (dosya bazlı).
        page_counts: Her kaynak dosyanın sayfa sayıları (per_document modu için).
    """
    import fitz

    doc = fitz.open(str(pdf_path))

    # Arial Bold font dosyası
    arial_bold = Path("C:/Windows/Fonts/arialbd.ttf")
    arial_font = Path("C:/Windows/Fonts/arial.ttf")
    if arial_bold.exists():
        fontfile = str(arial_bold)
        fontname = "arialbd"
    elif arial_font.exists():
        fontfile = str(arial_font)
        fontname = "arial"
    else:
        fontfile = None
        fontname = "helv"

    # Per-document modunda her sayfanın numarasını önceden hesapla
    if mode == "per_document" and page_counts:
        page_numbers: list[int] = []
        for doc_idx, count in enumerate(page_counts):
            page_numbers.extend([start + doc_idx] * count)
    else:
        page_numbers = [start + i for i in range(len(doc))]

    for i, page in enumerate(doc):
        number = page_numbers[i] if i < len(page_numbers) else start + i
        rect = page.rect
        text_rect = fitz.Rect(
            rect.width - 80, 8,
            rect.width - 10, 38
        )
        shape = page.new_shape()
        shape.insert_textbox(
            text_rect,
            str(number),
            fontname=fontname,
            fontfile=fontfile,
            fontsize=20,
            color=(1.0, 0.0, 0.0),
            align=fitz.TEXT_ALIGN_RIGHT,
        )
        shape.commit(overlay=True)

    doc.saveIncr()
    doc.close()
