"""
PDFToolKit - PDF'den Diğer Formatlara Dönüştürme

PDF → Word (DOCX), Excel (XLSX), JPG dönüştürme fonksiyonları.
"""

from pathlib import Path

from utils.logger import get_logger
from utils.file_utils import generate_output_filename, ensure_output_dir

logger = get_logger(__name__)


def pdf_to_word(
    pdf_path: str | Path,
    output_path: str | Path | None = None
) -> Path:
    """
    PDF dosyasını Word (DOCX) formatına dönüştürür.

    Args:
        pdf_path: Kaynak PDF dosyası.
        output_path: Çıktı dosya yolu. None ise otomatik üretilir.

    Returns:
        Oluşturulan DOCX dosyasının Path nesnesi.
    """
    from pdf2docx import Converter

    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {pdf_path}")

    # Çıktı yolu
    if output_path is None:
        output_path = generate_output_filename(pdf_path.stem, "", ".docx")
    else:
        output_path = Path(output_path)

    cv = Converter(str(pdf_path))
    cv.convert(str(output_path))
    cv.close()

    logger.info(f"PDF → Word: '{pdf_path.name}' → {output_path}")
    return output_path


def pdf_to_excel(
    pdf_path: str | Path,
    output_path: str | Path | None = None
) -> Path:
    """
    PDF dosyasındaki tabloları Excel (XLSX) formatına dönüştürür.

    Not: Sadece tablo içeren PDF'ler için doğru çalışır.
    Normal metin içerikli PDF'lerde sonuç sınırlı olabilir.

    Args:
        pdf_path: Kaynak PDF dosyası.
        output_path: Çıktı dosya yolu. None ise otomatik üretilir.

    Returns:
        Oluşturulan XLSX dosyasının Path nesnesi.
    """
    import fitz  # PyMuPDF
    from openpyxl import Workbook

    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {pdf_path}")

    # Çıktı yolu
    if output_path is None:
        output_path = generate_output_filename(pdf_path.stem, "", ".xlsx")
    else:
        output_path = Path(output_path)

    doc = fitz.open(str(pdf_path))
    wb = Workbook()

    for page_idx, page in enumerate(doc):
        if page_idx == 0:
            ws = wb.active
            ws.title = f"Sayfa {page_idx + 1}"
        else:
            ws = wb.create_sheet(title=f"Sayfa {page_idx + 1}")

        # Sayfadaki tabloları bul
        tables = page.find_tables()

        if tables and len(tables.tables) > 0:
            row_offset = 1
            for table in tables:
                data = table.extract()
                for r_idx, row in enumerate(data):
                    for c_idx, cell_value in enumerate(row):
                        ws.cell(
                            row=row_offset + r_idx,
                            column=c_idx + 1,
                            value=cell_value if cell_value else ""
                        )
                row_offset += len(data) + 1  # Tablolar arası boşluk
        else:
            # Tablo yoksa metni satır satır yaz
            text = page.get_text()
            for r_idx, line in enumerate(text.split("\n"), start=1):
                ws.cell(row=r_idx, column=1, value=line)

    doc.close()
    wb.save(str(output_path))

    logger.info(f"PDF → Excel: '{pdf_path.name}' → {output_path}")
    return output_path


def pdf_to_jpg(
    pdf_path: str | Path,
    output_dir: str | Path | None = None,
    dpi: int = 200
) -> list[Path]:
    """
    PDF'nin her sayfasını ayrı JPG dosyası olarak kaydeder.

    PyMuPDF (fitz) kullanır — Poppler kurulumu gerektirmez.

    Args:
        pdf_path: Kaynak PDF dosyası.
        output_dir: Çıktı dizini. None ise varsayılan çıktı dizini kullanılır.
        dpi: Çözünürlük (varsayılan: 200 DPI).

    Returns:
        Oluşturulan JPG dosyalarının Path listesi.
    """
    import fitz  # PyMuPDF

    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {pdf_path}")

    output_dir = ensure_output_dir(output_dir)

    doc = fitz.open(str(pdf_path))
    zoom = dpi / 72  # fitz varsayılan 72 DPI
    matrix = fitz.Matrix(zoom, zoom)

    created_files = []
    for i, page in enumerate(doc, start=1):
        pix = page.get_pixmap(matrix=matrix)
        output_file = output_dir / f"{pdf_path.stem}_sayfa{i}.jpg"
        pix.save(str(output_file))
        created_files.append(output_file)

    doc.close()

    logger.info(
        f"PDF → JPG: '{pdf_path.name}' → {len(created_files)} görsel → {output_dir}"
    )
    return created_files
